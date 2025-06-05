import base64
import logging
import os

from django.forms import model_to_dict
from openpyxl import Workbook
from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration
from csp.decorators import csp_exempt
from django.urls import reverse_lazy
from django.db.models.functions import Coalesce, TruncDay, TruncMonth, TruncYear
from django.db.models import Count, Case as MCase, When, Q, OuterRef, Exists
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render, reverse, redirect, get_object_or_404, Http404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
from django.contrib.gis.geos import fromstr
from django.contrib.gis.db.models.functions import Distance
from django.contrib.auth.views import LoginView

from django.contrib.auth import login, logout
from django.contrib import messages
from django.views import View
from django.views.generic import CreateView, FormView, ListView, UpdateView
from django.utils import timezone
from datetime import timedelta  # Add this import
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.views import PasswordChangeView

from .user_forms import (
    UserRegistrationForm,
    VerifyOtpFrom,
    UserRegistrationCompleteForm,
    ResendMobileVerificationOtpForm,
    ForgotPasswordForm,
    ChangePasswordForm,
)

from .mixins import AdminRequiredMixin

from api.models import (
    Case,
    cUser,
    Like,
    Comment,
    CaseHistory,
    Victim,
    Criminal,
    PoliceStation,
    PoliceOfficer,
    UserOTPBaseKey,
    Media,
    LostVehicle,
)

# from rest_framework.request import Request
from api.view.district_views import *
from api.view.police_views import *
from api.view.case_view import *
from api.view.cuser_views import *
from api.forms.user import AddOfficerForm, RemoveOfficerForm, ChangeDesignationForm
from api.forms.case import CaseForm, CaseUpdateForm
from api.utils import get_cases, detect_malicious_patterns_in_media_fileobj
import magic
import tempfile
import mimetypes
import re

logger = logging.getLogger(__name__)

def information(request):
    logger.info("Entering information function") 
    """
    limit = int(request.GET.get("limit", 10))
    page = int(request.GET.get("page", 0))
    cases = Case.objects.all()
    items = []
    for case in cases:
        number of comments = case.comment_set.count()

        case = {
            "number": case.cid,
            "type": dict(Case.cType)[case.type],
            "status": dict(Case.cState)[case.cstate],
            "likes": 10,
            "comments": number_of_comments,
            "reported_date": str(case.created),
        }
        items.append(case)

    vars = {"items": items, "home": True}
    logger.info("Exiting information function")
    """
    information = Information.objects.all()
    logger.info(" Exiting information function") 
    return render(request, "information.html", {'informations':information})


from django.shortcuts import render
from .models import Emergency, EmergencyType
from django.contrib.gis.geos import Point
from django.contrib.gis.db.models.functions import Distance
from django.core.exceptions import ValidationError
import logging

logger = logging.getLogger(__name__)

def emergency(request):
    logger.info("Entering emergency function")
    user_lat = request.GET.get("lat")
    user_long = request.GET.get("long")
    selected_emergency_type = request.GET.get("emergency_type")

    logger.debug(
        f"Request parameters - lat: {user_lat}, long: {user_long}, type: {selected_emergency_type}"
    )

    emergencies = Emergency.objects.all()

    # Filter by location if provided
    if user_lat and user_long:
        try:
            user_location = Point(float(user_long), float(user_lat), srid=4326)
            emergencies = emergencies.annotate(
                distance=Distance("geo_location", user_location)
            ).order_by("distance")
        except (ValueError, TypeError) as e:
            logger.error(f"Error processing coordinates: {str(e)}")
            # Optionally: raise ValidationError("Invalid coordinates provided")

    # Filter by emergency type if provided
    if selected_emergency_type:
        emergencies = emergencies.filter(tid_id=selected_emergency_type)

    emergency_types = EmergencyType.objects.all()
    logger.info(f" Emergencies :%s",list(emergencies.values())if hasattr(emergencies,"values") else list(emergencies))

    logger.info("Rendering full emergency template")
    return render(
        request,
        "emergency.html",
        {
            "items": emergencies,
            "emergency_types": emergency_types,
            "selected_emergency_type": selected_emergency_type,
            "user_lat": user_lat,
            "user_long": user_long,
        },
    )

def logout_view(request):
    logger.info("Entering logout_view function")
    logout(request)
    logger.info("Exiting logout_view function")
    return redirect("/")  # Replace 'home' with the URL name of your home page
    #return HttpResponseRedirect("/")  # Replace 'home' with the URL name of your home page


class HomePageView(LoginRequiredMixin, View):
    """
    View for displaying case listings with authentication required.
    Inherits from LoginRequiredMixin to enforce authentication.
    """
    login_url = '/login/'  # Redirect URL for unauthenticated users
    redirect_field_name = 'next'  # GET parameter name for the redirect URL
    paginate_by = 10  # Number of items per page

    def get_header(self):
        logger.info("Entering get_header")
        header_map = {
            "my-complaints": "My complaints",
            "stolen_vechicle": "Stolen Vehicle",
            "drug_case": "Drug Case Reported",
            "extortion_case": "Extortion Case Reported",
        }
        return header_map.get(self.kwargs.get("title"), "")

    def get_case_type(self):
        logger.info("Entering get_case_type")
        case_type_mapping = {
            "stolen-vehicle": "vehicle",
            "drug-case": "drug",
            "extortion-case": "extortion",
            "vehicle": "vehicle",
            "case": "drug",
            "case": "extortion",
        }
        case_type_from_url = self.kwargs.get("case_type")
        logger.info(f" Case type from url : {case_type_from_url}")
        mapped_type = case_type_mapping.get(case_type_from_url)
        logger.info(f" Exiting get_case_type with mapped_type : {mapped_type}")
        return mapped_type

    def get_template_name(self):
        logger.info("Entering get_template_name")
        template = self.get_case_type()
        if template:
            logger.info(f"Exiting get_template_name with template: {template}")
            return f"case/{template}.html"
        logger.info("Exiting get_template_name with default home.html")
        return "home.html"

    def get_queryset(self):
        logger.info("Entering get_queryset in HomePageView")

        user = self.request.user

        logger.info(f"Total Case in Records in DB :{Case.objects.count()}")

        # Start with base queryset and annotate as needed
        """
        cases = (
            Case.objects.annotate(
                comment_count=Count("comment", distinct=True),
                like_count=Count("likes", distinct=True),
                is_location_visible=MCase(
                    When(
                        cstate="pending",
                        pid__policeofficer__user_id=user.id,
                        then=True,
                    ),
                    default=False,
                ),
            )
            .select_related("pid", "oid", "oid__user", "oid__pid", "oid__pid__did")
            .order_by("-created")
        )"""
        cases = Case.objects.all()

        logger.info(f"Total Case initial count :{cases.count()}")

        # Handle search functionality (only this stays in the view)
        if q := self.request.GET.get("q"):
            search_filter = (
                Q(lostvehicle__regNumber=q)
                | Q(lostvehicle__chasisNumber=q)
                | Q(lostvehicle__engineNumber=q)
            )
            try:
                int(q)
                search_filter |= Q(cid=q)
            except ValueError:
                pass
            cases = cases.filter(search_filter)
        logger.info(f"Total Case after search :{cases.count()}")

        # Use helper for all other filtering
        #case_type = self.get_case_type()
        case_type = self.request.GET.get("type")
        logger.info(f" CASE TYPE REQUEST RECEIVED IS :{case_type}")

        my_complaints = self.kwargs.get("case_type") == "my-complaints"
        logger.info(f"Total Case finally returning :{cases.count()}")
        ret_cases = get_cases(user, cases, case_type=case_type, my_complaints=my_complaints)
        logger.info(f"Total Case finally returning :{ret_cases.count()}")

        return ret_cases.order_by("-created").distinct()

    def get(self, request, *unused_args, **unused_kwargs):
        """
        Handle GET requests for the home page, including pagination and filtering.
        """
        logger.info("Entering get with request: %r", request)
        if request.GET.get("action") == "logout":
            logout(request)
            return redirect(reverse("login"))
        cases = self.get_queryset()
        header = self.get_header()
        template_name = self.get_template_name()

        # Pagination with input validation to prevent cyber attacks
        page_number = request.GET.get('page', 1)
        try:
            page_number = int(page_number)
            if page_number < 1:
                page_number = 1
        except (ValueError, TypeError):
            page_number = 1

        paginator = Paginator(cases, self.paginate_by)
        try:
            page_obj = paginator.get_page(page_number)
        except (AttributeError, ValueError):
            logger.error("Pagination error: invalid page number %r", page_number)
            page_obj = paginator.get_page(1)

        # Get page range for pagination controls
        page_range = self.get_page_range(paginator, page_obj)

        context = {
            "cases": page_obj,
            "header": header,
            "page_obj": page_obj,
            "page_range": page_range,
            "is_paginated": page_obj.has_other_pages(),
        }

        logger.info(
            "Exiting get with template: %s, page %s, total pages: %s",
            template_name,
            page_number,
            paginator.num_pages,
        )
        return render(
            request,
            template_name,
            context,
        )
    
    def get_page_range(self, paginator, page_obj):
        """Return a range of page numbers to display in pagination controls."""
        page_number = page_obj.number
        total_pages = paginator.num_pages
        
        # Always show first, last, and 5 pages around current page
        if total_pages <= 7:
            # If there are 7 or fewer pages, show all
            return range(1, total_pages + 1)
        
        # Complex case: need to add ellipsis
        if page_number <= 4:
            # Near the start
            return list(range(1, 6)) + ["..."] + [total_pages]
        elif page_number >= total_pages - 3:
            # Near the end
            return [1] + ["..."] + list(range(total_pages - 4, total_pages + 1))
        else:
            # In the middle
            return [1] + ["..."] + list(range(page_number - 2, page_number + 3)) + ["..."] + [total_pages]


class UserRegistrationView(View):
    def get(self, request, *args, **kwargs):
        logger.info(f"Entering get with request: {request}")
        form = UserRegistrationForm
        # with open("/cfa_server/api.log","a") as file:
        #   file.write(" Entering user sign up ")
        return render(request, "api/signup.html", {"form": form})

    def post(self, request, *args, **kwargs):
        logger.info(f"Entering post with request: {request}")
        form = UserRegistrationForm(request.POST)
        #import ipdb

        #ipdb.set_trace()
        if form.is_valid():
            user = form.save()
            UserOTPBaseKey.send_otp_verification_code(user)
            request.session["mobile"] = user.mobile
            return redirect(reverse("verify_mobile") + f"?mobile={user.mobile}")
        else:
            mobile = request.POST.get("mobile")
            try:
                user = cUser.objects.get(mobile=mobile)
            except cUser.DoesNotExist:
                pass
            else:
                return redirect(
                    reverse("verify_mobile") + f"?mobile={request.POST['mobile']}"
                )
                # if user.is_verified:
                #     messages.info(
                #         request,
                #         "A user with this mobile is already registered. "
                #         "If this is your mobile please login.",
                #     )
                # else:
                #     request.session["mobile"] = user.mobile
                #     messages.info(
                #         request,
                #         "A user with this mobile is already registered. "
                #         "But the mobile is not verified yet. If this is"
                #         " your mobile please verify it. You can resend verification"
                #         " otp from <a href='/accounts/resend-verification-otp/'>"
                #         "here</a>.",
                #     )
        return render(request, "api/signup.html", {"form": form})


class ResendMobileVerificationOtpView(FormView):
    template_name = "api/resend_mobile_verificatio_otp.html"
    form_class = ResendMobileVerificationOtpForm
    success_url = reverse_lazy("verify_mobile")

    def form_valid(self, form):
        user = form.save()
        if user:
            self.request.session["mobile"] = user.mobile
        return HttpResponseRedirect(self.get_success_url())


class UserRegistrationCompleteView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        form = UserRegistrationCompleteForm
        return render(request, "api/signup_complete.html", {"form": form})

    def post(self, request, *args, **kwargs):
        form = UserRegistrationCompleteForm(
            data=request.POST,
            files=request.FILES,
            instance=request.user,
            request=request,
        )
        if form.is_valid():
            user = form.save()
            # UserOTPBaseKey.send_otp_verification_code(user)
            # login(request, user)
            return redirect(reverse("login"))
        return render(request, "api/signup_complete.html", {"form": form})


class VerifyOtpView(View):
    def get(self, request, *args, **kwargs):
        form = VerifyOtpFrom(mobile=request.session.get("mobile"))
        return render(request, "api/verify_otp.html", {"form": form})

    def post(self, request, *args, **kwargs):
        mobile = request.session.get("mobile")
        form = VerifyOtpFrom(request.POST, mobile=mobile)
        if form.is_valid():
            user = form.save()
            messages.info(
                request,
                "Mobile verification successful."
                " Complete your registration and add your password now.",
            )
            try:
                del request.session["mobile"]
            except KeyError:
                pass
            login(request, user, "axes.backends.AxesBackend")
            return redirect(reverse("complete_signup"))
        return render(request, "api/verify_otp.html", {"form": form})


class CaseAddView(LoginRequiredMixin, CreateView):
    form_class = CaseForm
    model = Case
    template_name = "case/add_case.html"
    success_url = "/"

    @csp_exempt
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["user"] = self.request.user
        return kw

    def form_valid(self, form):
        form.save()
        return redirect(reverse("home"))


class AddCommentView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return redirect("/")

    def post(self, request, *args, **kwargs):
        data = request.POST
        case_id = kwargs["case_id"]
        comment = data["comment"]
        Comment.objects.create(cid_id=case_id, content=comment, user=request.user)
        return redirect("/")


class AddLikeView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        logger.info("Entering get with request: {request}")
        case_id = kwargs["case_id"]
        Like.objects.get_or_create(case_id=case_id, user=request.user)
        messages.success(request, "Your like is added.")
        return redirect("/")


class ChangeCaseStateUpdateView(LoginRequiredMixin, UpdateView):
    template_name = "case/case_update.html"
    form_class = CaseUpdateForm
    queryset = Case.objects.all()
    success_url = reverse_lazy("home")

    def get_queryset(self):
        logger.info("Entering get_queryset")
        queryset = self.queryset.filter(
            (Q(cstate="pending") & Q(pid__policeofficer__user=self.request.user))
            | Q(oid__user=self.request.user)
        ).distinct()
        return queryset

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["request"] = self.request
        return kw


class GetCaseHistory(View):
    def get(self, request, *args, **kwargs):
        case_id = kwargs["case_id"]
        case_histories = (
            CaseHistory.objects.filter(case_id=case_id)
            #.prefetch_related("medias") # Need to add the result of medias if any 
            .order_by("created")
        )
        from itertools import pairwise

        history_pairs = list(pairwise(case_histories))
        context = {"history_pairs": history_pairs}
        html = render_to_string(
            "case/case_history_json.html", request=request, context=context
        )
        return JsonResponse({"html": html})


class CrimeListView(ListView):
    def get_queryset(self):
        logger.info("Entering get_queryset")
        model = self.get_model()
        qs = model.objects.all()
        crime_type = self.get_crime_type()
        if crime_type:
            qs = qs.filter(type=crime_type)
        return qs

    def get_crime_type(self):
        crime_type = {
            "missing-children": "missing_children",
            "children-found": "children_found",
            "missing-person": "missing_person",
            "dead-body": "dead_body",
            "other": "other",
            "offender": "offender",
            "wanted": "wanted",
            "proclaimed": "proclaimed",
        }
        return crime_type.get(self.kwargs["crime_type"])

    def get_model(self):
        logger.info("Entering get_model")
        crime_type = self.get_crime_type()
        victims = [
            "missing_children",
            "children_found",
            "missing_person",
            "dead_body",
            "other",
        ]
        if crime_type is None or crime_type in victims:
            return Victim
        else:
            return Criminal

    def get_template_names(self):
        logger.info("Entering get_template_names")
        crime_type = self.get_crime_type() or "missing_children"
        return [f"case/{crime_type}.html"]


class ForgotPasswordView(FormView):
    form_class = ForgotPasswordForm
    template_name = "users/forgot_password.html"
    success_url = reverse_lazy("reset_password_web")

    def form_valid(self, form):
        form.save()
        self.request.session["mobile"] = form.cleaned_data["mobile"]
        self.request.session["password_reset"] = True
        return super().form_valid(form)


class ResetPasswordView(FormView):
    form_class = ChangePasswordForm
    template_name = "users/reset_password.html"
    success_url = reverse_lazy("login")

    def dispatch(self, request, *args, **kwargs):
        if not self.request.session.get("password_reset"):
            raise Http404
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.save()
        messages.info(self.request, "Password change successful.")
        return super().form_valid(form)

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["request"] = self.request
        return kw


class NearestPoliceStationsView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        logger.info("Entering get nearest police staiton view")
        lat = request.GET.get("lat")
        long = request.GET.get("long")
        if lat and long:
            location = fromstr(f"POINT({self.long} {self.lat})", srid=4326)
            user_distance = Distance("geo_location", location)
            qs = PoliceStation.objects.annotate(radius=user_distance).order_by("radius")
        else:
            qs = PoliceStation.objects.all()
        qs = qs.values("pid", "did", "name", "address")
        return JsonResponse(list(qs), safe=False)

from api.utils import get_cases, generate_pdf_from_cases,build_case_list_with_details
class ExportCrime(View):
    def get(self, request, *args, **kwargs):
        logger.info(f"Exporting crime for user: {self.request.user}")
        cases = get_cases(request.user)
        return self.get_pdf(cases)

    def get_pdf(self, cases):
        logger.info("Entering get_pdf")
        font_config = FontConfiguration()
        template_name = "export/pdf.html"
        user = self.request.user
        # Generate detailed data for PDF from cases
        cases = get_cases(user)
        cases = cases.order_by("created")
        #data = [model_to_dict(case) for case in cases]  # or however you serialize your case objects
        data = []
        for case in cases:
            d = model_to_dict(case)
            d["created"] = case.created
            data.append(d)
        detailed_data = build_case_list_with_details(data)
        
        # Generate PDF file and get the output path.
        # generate_pdf_from_cases() should write the file and return the output path.
        from django.utils import timezone
        output_filename = f"{user.mobile}_{timezone.now().strftime('%Y%m%d_%H%M%S')}_testReport.pdf"
        output_dir = os.path.join("media", "files")
        os.makedirs(output_dir, exist_ok=True)
        output_path = generate_pdf_from_cases(detailed_data, os.path.join(output_dir, output_filename))
        
        # Return the PDF file as an attachment for download.
        absolute_path = os.path.abspath(output_path)
        logger.info("PDF absolute path: %s", absolute_path)
        filename = os.path.basename(output_path)
        with open(absolute_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PoliceStationListView(AdminRequiredMixin, ListView):
    template_name = "stations/list.html"
    model = PoliceStation
    queryset = PoliceStation.objects.all()

    def get_queryset(self):
        logger.info("Entering get_queryset in police station list view")
        qs = super().get_queryset()
        if q := self.request.GET.get("q"):
            qs = qs.filter(Q(name__icontains=q) | Q(address__icontains=q))
        return qs


class ChoosePoliceOfficerView(AdminRequiredMixin, ListView):
    template_name = "stations/choose_police_officers.html"
    model = cUser
    queryset = cUser.objects.filter(is_superuser=False)

    def get_queryset(self):
        qs = super().get_queryset()
        if q := self.request.GET.get("q"):
            qs = qs.filter(mobile__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["station"] = get_object_or_404(
            PoliceStation, pk=self.kwargs["station_id"]
        )
        context["add_officer_form"] = AddOfficerForm
        return context


class PoliceOfficerListView(AdminRequiredMixin, ListView):
    template_name = "stations/police_officers.html"
    model = PoliceOfficer

    def get_queryset(self):
        logger.info("Entering get_queryset police officer list view ")
        qs = super().get_queryset()
        qs = qs.filter(pid_id=self.kwargs["station_id"])
        if q := self.request.GET.get("q"):
            qs = qs.filter(user__mobile__icontains=q)
        return qs.select_related("user")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["station"] = get_object_or_404(
            PoliceStation, pk=self.kwargs["station_id"]
        )
        context["remove_officer_form"] = RemoveOfficerForm
        return context


class AddOfficerView(View):
    def post(self, request, *args, **kwargs):
        logger.info("Entering post in addOfficerView")
        user_id = int(request.POST.get("user"))
        user = cUser.objects.get(pk=user_id)
        user.role = "police"
        user.save()
        pid = get_object_or_404(PoliceStation, pk=kwargs["station_id"])
        rank = request.POST.get("rank")
        PoliceOfficer.objects.update_or_create(
            user=user,
            defaults={
                "pid": pid,
                "rank": rank,
                "mobile": user.mobile,
            },
        )
        messages.success(request, "Succesfully added user to police list.")
        return redirect(reverse("police_officer_list", args=[pid.pk]))


class RemoveOfficerView(View):
    def post(self, request, *args, **kwargs):
        logger.info("Entering post in RemoveOfficerView")
        user_id = int(request.POST.get("user"))
        user = cUser.objects.get(pk=user_id)
        user.role = "user"
        user.save()
        user.policeofficer_set.all().delete()
        messages.success(request, "Succesfully removed user from officer list.")
        return redirect(reverse("police_officer_list", args=[kwargs["station_id"]]))


class RemovePoliceOfficerListView(AdminRequiredMixin, ListView):
    template_name = "stations/choose_police_officers_to_remove.html"
    model = PoliceOfficer

    def get_queryset(self):
        logger.info("Entering get_queryset RemovePoliceOfficerListView")
        qs = super().get_queryset()
        qs = qs.filter(pid_id=self.kwargs["station_id"])
        if q := self.request.GET.get("q"):
            qs = qs.filter(mobile__icontains=q)
        return qs.select_related("user")

    def get_context_data(self, **kwargs):
        logger.info("Entering get_context_data RemovePoliceOfficerListView")
        context = super().get_context_data(**kwargs)
        context["station"] = get_object_or_404(
            PoliceStation, pk=self.kwargs["station_id"]
        )
        context["remove_officer_form"] = RemoveOfficerForm
        return context


class AssignDesignationListView(AdminRequiredMixin, ListView):
    template_name = "stations/police_designation_list.html"
    model = cUser
    queryset = cUser.objects.all()


class ChangeDesignation(AdminRequiredMixin, FormView):
    form_class = ChangeDesignationForm
    template_name = "stations/change_designation.html"

    def form_valid(self, form):
        form.save()
        return redirect(reverse("assign_designation_list"))

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["user"] = get_object_or_404(cUser, pk=self.kwargs["user_id"])
        return kw


class CustomLoginView(LoginView):
    template_name = "admin/login.html"  # Use your custom login template


class CustomAdminPasswordChangeForm(PasswordChangeForm):
    template_name = "admin/custom_password_change.html"


class CustomPasswordChangeView(PasswordChangeView):
    # Customize the view as needed

    form_class = CustomAdminPasswordChangeForm
    template_name = (
        "admin/custom_password_change.html"  # Replace with your custom template path
    )
    success_url = reverse_lazy(
        "admin:index"
    )  # Redirect to the admin index page after password change


def custom_404_view(request, exception):
    return render(request, "custom_error.html", status=404)


def custom_400_view(request, exception=None):
    return render(request, "custom_error.html", status=400)


def custom_401_view(request, exception=None):
    return render(request, "custom_error.html", status=401)


def custom_403_view(request, exception=None):
    return render(request, "custom_error.html", status=403)


from django.shortcuts import render
from .models import AboutPage


def about(request):
    logger.info("Entering about function")
    about_content = (
        AboutPage.objects.first()
    )  # Get the first (and only) AboutPage object
    logger.info("Exiting about function")
    return render(request, "about.html", {"about_content": about_content})


from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from .models import Case
from django.db.models import Count

class dboardView(View):
    def get(self, request, *arg, **kwargs):
        logger.debug(f"DASHBOARDH GET ENTERING : {request}")
        return(request, "dashboard.html")

@login_required
def dashboard(request):
    logger.info(f"Entering dashboard function with request: {request}")
    if not request.user.is_authenticated:
        return redirect('login')
    
    user = request.user
    ctype = request.GET.get('ctype', None)

    # Use get_cases utility for all filtering
    cases = get_cases(user, case_type=ctype)

    case_count = cases.count()
    drug_count = cases.filter(type="drug").count()
    extortion_count = cases.filter(type="extortion").count()
    vehicle_count= cases.filter(type="vehicle").count()

    # Overall case summary by type
    case_summary = cases.values('type').annotate(
        count=Count('cid')
    ).order_by('type')
    
    # Status summaries by case type (filtered by user's permissions)
    drug_status_summary = cases.filter(type='drug').values('cstate').annotate(
        count=Count('cid')
    ).order_by('cstate')
    
    extortion_status_summary = cases.filter(type='extortion').values('cstate').annotate(
        count=Count('cid')
    ).order_by('cstate')
    
    vehicle_status_summary = cases.filter(type='vehicle').values('cstate').annotate(
        count=Count('cid')
    ).order_by('cstate')
    
    # Generate time-based data using Trunc functions without time restrictions
    daily_data = (
        cases
        .annotate(date=TruncDay('created'))
        .values('date')
        .annotate(count=Count('cid'))
        .order_by('date')
    )
    daily_data_list = [
        {'date': entry['date'].strftime('%Y-%m-%d'), 'count': entry['count']}
        for entry in daily_data if entry['date']
    ]
    
    monthly_data = (
        cases
        .annotate(date=TruncMonth('created'))
        .values('date')
        .annotate(count=Count('cid'))
        .order_by('date')
    )
    monthly_data_list = [
        {'date': entry['date'].strftime('%b %Y'), 'count': entry['count']}
        for entry in monthly_data if entry['date']
    ]
    
    yearly_data = (
        cases
        .annotate(date=TruncYear('created'))
        .values('date')
        .annotate(count=Count('cid'))
        .order_by('date')
    )
    yearly_data_list = [
        {'date': entry['date'].strftime('%Y'), 'count': entry['count']}
        for entry in yearly_data if entry['date']
    ]
    
    time_data = {
        'daily': daily_data_list,
        'monthly': monthly_data_list,
        'yearly': yearly_data_list
    }
    
    colors = {
        'pending': '#FF9800',
        'accepted': '#4CAF50',
        'found': '#2196F3',
        'assign': '#9C27B0',
        'visited': '#FF5722',
        'inprogress': '#03A9F4',
        'transfer': '#795548',
        'resolved': '#8BC34A',
        'info': '#607D8B',
        'rejected': '#F44336'
    }
    
    context = {
        'case_summary': case_summary,
        'case_summary_total': case_count,
        'ctype_filter': ctype,
        'drug_status_summary': drug_status_summary,
        'drug_status_summary_total': drug_count,
        'extortion_status_summary': extortion_status_summary,
        'extortion_status_summary_total': extortion_count,
        'vehicle_status_summary': vehicle_status_summary,
        'vehicle_status_summary_total': vehicle_count,
        'colors': colors,
        'time_data': time_data,
    }
    
    logger.info("Exiting dashboard function")
    return render(request, 'dashboard.html', context)
    #return render(request, 'dash.html', context)
    
    from django.http import JsonResponse
from django.template.loader import render_to_string

def get_case_comments(request, cid):
    logger.info(f"Entering get_case_comment fpr cid :{cid}")
    # Fetch comments for the case
    case = Case.objects.filter(cid=cid).first()
    if case is None:
        logger.info(f" No record found. Exiting get_case_comment")
        return JsonResponse({'html': ''})
    comments = Comment.objects.filter(cid=case).order_by('-created')
    html = render_to_string('case/comments_partial.html', {'comments': comments})
    logger.info(f" Returning {comments.count} comment records for cid :{cid}")
    return JsonResponse({'html': html})

def get_case_history(request, cid):
    logger.info(f" Entering get_case_history for cid :{cid}")
    # Fetch history for the case
    case = Case.objects.filter(cid=cid).first()
    if case is None:
        logger.info("No record found. Exiting get_case_history ")
        return JsonResponse({'html': ''})
    history = CaseHistory.objects.filter(case=case).order_by('-created')
    html = render_to_string('case/history_partial.html', {'history': history})
    logger.info(f"Returning {history.count()} history records for cid {cid}")
    return JsonResponse({'html': html})
    
def append_case_medias(page_obj):
    """
    For each case in page_obj, attach a .medias attribute with the related Medias
    where parentId=case.cid and source='case'.
    """
    case_ids = [case.cid for case in page_obj]
    # Fetch all relevant medias in one query
    medias = Media.objects.filter(parentId__in=case_ids, source="case")
    medias_by_case = {}
    for media in medias:
        medias_by_case.setdefault(media.parentId, []).append(media)
    # Attach medias to each case
    for case in page_obj:
        case.medias = medias_by_case.get(case.cid, [])
    return page_obj

def get_media(request):
    source = request.GET.get('source')
    cid = request.GET.get('cid')
    logger.info(f"Entering get_media with source: {source}, cid: {cid}")
    if not source or not cid:
        return JsonResponse({'html': '<div class="text-danger">Missing parameters.</div>'})
    medias = Media.objects.filter(parentId=cid, source=source)
    html = render_to_string('case/media_partial.html', {'medias': medias})
    logger.info(f"Returning {medias.count()} media records for source: {source}, cid: {cid}")
    return JsonResponse({'html': html})

def get_case_media(request, cid):
    logger.info(f"Entering get_case_media with cid: {cid}")
    medias = Media.objects.filter(parentId=cid, source='case')
    html = render_to_string('case/case_media_partial.html', {'medias': medias})
    logger.info(f"Exiting get_cse_media with: {medias.count()} records")
    return JsonResponse({'html': html})

from django.contrib.auth.decorators import login_required
from django.http import FileResponse, Http404
from django.conf import settings
import os
@login_required
def protected_media(request, path):
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'))
    raise Http404()


import openpyxl
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
import os

class UploadLostVehicleView(LoginRequiredMixin, View):
    template_name = "uploadlostvehicle.html"

    def get(self, request):
        logger.info("UploadLostVehicleView GET: Rendering upload form for user %s", request.user)
        return render(request, self.template_name)

    def post(self, request):
        errors = []
        logger.info("UploadLostVehicleView POST: Received file upload request from user %s", request.user)
        xls_file = request.FILES.get("xls_file")
        if not xls_file:
            logger.warning("No file uploaded in request by user %s", request.user)
            errors.append("No file uploaded.")
            return render(request, self.template_name, {"errors": errors})

        # Use detect_malicious_patterns_in_media_fileobj to check file type, size, and security
        if detect_malicious_patterns_in_media_fileobj(xls_file, xls_file.name):
            logger.warning("Suspicious or disallowed Excel file uploaded by user %s", request.user)
            errors.append("Suspicious, disallowed, or too large file detected. Upload rejected.")
            return render(request, self.template_name, {"errors": errors})

        # Save to temp file for openpyxl
        xls_file.seek(0)
        xls_content = xls_file.read()
        xls_file.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(xls_content)
            file_path = tmp.name
        logger.info("Saved uploaded file to temp file: %s", file_path)

        try:
            logger.info("Opening uploaded file with openpyxl: %s", file_path)
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            # Read only the first 8 columns from the header
            header = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=8))]
            header_lower = [h.lower() for h in header]
            rows = list(sheet.iter_rows(min_row=2, max_col=8, values_only=True))
            user = request.user
            ouser = cUser.objects.filter(mobile=user.mobile).first()
            logger.info("Resolved cUser for upload: %s", ouser)

            if not ouser:
                logger.error("User %s not found in cUser table.", user)
                errors.append(f"{user}: You are not found in the system.")
                return render(request, self.template_name, {"errors": errors})

            police_officer = PoliceOfficer.objects.filter(user=ouser).first()
            if not police_officer:
                logger.error("User %s has no associated PoliceStation.", user)
                errors.append(f"{user}: You have no associated PoliceStation.")
                return render(request, self.template_name, {"errors": errors})
            pid = police_officer.pid

            officer = PoliceOfficer.objects.filter(pid=pid, report_on_this=True).first()
            if not officer:
                logger.error("No reporting officer found for station %s (user %s)", pid.name, user)
                errors.append(f"{user}: No reporting officer found for station {pid.name}.")
                return render(request, self.template_name, {"errors": errors})

            lat = pid.lat
            long = pid.long

            logger.info("Processing %d rows from uploaded file for user %s", len(rows), user)
            for row_idx, row in enumerate(rows, start=2):
                # Only use the first 8 columns
                row = row[:8]
                if not any(row):
                    logger.warning("Row %d is empty, skipping.", row_idx)
                    continue
                row_data = {header_lower[i]: (row[i] if i < len(row) and row[i] is not None else "") for i in range(len(header_lower))}
                reg_number = str(row_data.get("regnumber", "")).strip()
                description = str(row_data.get("description", "")).strip()
                if not reg_number or not description:
                    if not reg_number and not description:
                        logger.warning(
                            "Row %d: Missing both regNumber and description.",
                            row_idx
                        )
                        errors.append(
                            f"Row {row_idx}: Both regNumber and description are missing."
                        )
                    elif not reg_number:
                        logger.warning(
                            "Row %d: Missing regNumber for description: %s.",
                            row_idx,
                            description
                        )
                        errors.append(
                            f"Row {row_idx}: regNumber is missing for description: "
                            f"{description if description else '[empty]'}."
                        )
                    elif not description:
                        logger.warning(
                            "Row %d: Missing description for regNumber: %s.",
                            row_idx,
                            reg_number
                        )
                        errors.append(
                            f"Row {row_idx}: Description is missing for regNumber: "
                            f"{reg_number if reg_number else '[empty]'}."
                        )
                    continue

                if LostVehicle.objects.filter(regNumber=reg_number).exists():
                    logger.info("Row %d: regNumber %s already exists in LostVehicle.", row_idx, reg_number)
                    errors.append(f"Row {row_idx}: regNumber {reg_number} already exists.")
                    continue

                case = Case.objects.create(
                    user=user,
                    pid=pid,
                    oid=officer,
                    type="vehicle",
                    cstate="pending",
                    lat=lat,
                    long=long,
                    description=description,
                )
                logger.info("Created Case %s for regNumber %s", case.cid, reg_number)

                def safe_text(val, maxlen=100):
                    val = re.sub(r'[^\w\s\-.,]', '', str(val))
                    return val[:maxlen]

                chassis_number = safe_text(row_data.get("chassisnumber", ""), 50)
                engine_number = safe_text(row_data.get("enginenumber", ""), 50)
                make = safe_text(row_data.get("make", ""), 100)
                model_name = safe_text(row_data.get("model", ""), 100)            
                vehicle_lost_type = safe_text(row_data.get("losttype", "") or "abandoned", 30)
                if vehicle_lost_type not in ["abandoned", "stolen"]:
                    logger.warning("Row %d: Invalid vehicle lost type '%s', defaulting to 'abandoned'.", row_idx, vehicle_lost_type)
                    vehicle_lost_type = "abandoned"
                status = str(row_data.get("status", "")).strip().lower()

                lv = LostVehicle.objects.create(
                    caseId=case,
                    regNumber=reg_number,
                    description=description,
                    chasisNumber=chassis_number,
                    engineNumber=engine_number,
                    make=make,
                    model=model_name,
                    color="",  # No color column in the 8-column spec
                    vehicle_lost_type=vehicle_lost_type,
                )
                logger.info("Created LostVehicle %s for regNumber %s", lv.pk, reg_number)
                CaseHistory.objects.create(
                    case=case,
                    user=user,
                    cstate=case.cstate,
                    description="Records uploaded by police officer"
                )
                logger.info("Case history for case %s added", case.cid)

                if status == "found":
                    case.cstate = status
                    case.save()
                    logger.info("Updated case %s status to %s", case.cid, status)
                    CaseHistory.objects.create(
                        case=case,
                        user=user,
                        cstate=case.cstate,
                        description=description
                    )
                    logger.info("Case history for status update for case %s added", case.cid)

            if errors:
                logger.warning("Upload completed with errors for user %s: %s", user, errors)
                return render(request, self.template_name, {
                    "errors": [f"Following vehicles were not added. Add them separately: {', '.join(errors)}"]
                })
            else:
                logger.info("Upload completed successfully for user %s", user)
                return render(request, self.template_name, {"success": "All records uploaded successfully."})

        except Exception as e:
            logger.exception("Failed to process file upload for user %s: %s", request.user, e)
            errors.append(f"Failed to process file: {e}")
            return render(request, self.template_name, {"errors": errors})
        finally:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    logger.info("Temporary file %s removed after processing.", file_path)
                except Exception as cleanup_exc:
                    logger.warning("Failed to remove temporary file %s: %s", file_path, cleanup_exc)